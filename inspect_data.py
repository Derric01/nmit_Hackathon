import openpyxl

wb = openpyxl.load_workbook('Copy of SmartCampus_OneSheet_Dataset.xlsx')
print('Sheet names:', wb.sheetnames)
ws = wb.active
print('Active sheet:', ws.title)
print('Dimensions:', ws.dimensions)
print('Max row:', ws.max_row, 'Max col:', ws.max_column)
print()
print('--- Headers (Row 1) ---')
headers = [cell.value for cell in ws[1]]
print(headers)
print()
print('--- First 10 rows ---')
for row in ws.iter_rows(min_row=2, max_row=11, values_only=True):
    print(list(row))
print()
print('--- Last 3 rows ---')
for row in ws.iter_rows(min_row=ws.max_row-2, max_row=ws.max_row, values_only=True):
    print(list(row))
